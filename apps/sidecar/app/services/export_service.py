"""Export local wordbook (vocab + sentences) to Excel/PDF."""

from __future__ import annotations

from datetime import datetime, timezone
from io import BytesIO
from pathlib import Path
from urllib.parse import quote
from xml.sax.saxutils import escape

from app.services.tokenizer_service import tokenize_ja
from app.services.vocab_service import get_all_sentences_flat, get_all_vocab_items


def _playback_columns(pb: dict | None) -> tuple[str, str, str, str, str, str]:
    if not pb:
        return ("", "", "", "", "", "")
    return (
        str(pb.get("platform") or ""),
        str(pb.get("url") or ""),
        str(pb.get("title") or ""),
        str(pb.get("series_name") or ""),
        str(pb.get("episode_name") or ""),
        "" if pb.get("current_time") is None else str(pb.get("current_time")),
    )


def _normalize_export_dt(value: str | None, *, end: bool = False) -> str:
    text = str(value or "").strip()
    if not text:
        return ""
    if len(text) == 10:
        text = f"{text}T{'23:59:59.999999' if end else '00:00:00'}"
    if text.endswith("Z"):
        text = text[:-1] + "+00:00"
    try:
        dt = datetime.fromisoformat(text)
    except ValueError:
        return ""
    if dt.tzinfo is None:
        return dt.astimezone(timezone.utc).isoformat()
    return dt.astimezone(timezone.utc).isoformat()


def normalize_export_range(start: str | None = None, end: str | None = None) -> tuple[str, str]:
    return _normalize_export_dt(start), _normalize_export_dt(end, end=True)


def export_filename(ext: str, start_at: str = "", end_at: str = "") -> str:
    stamp = datetime.now(timezone.utc).strftime("%Y%m%d")
    suffix = ""
    if start_at or end_at:
        a = (start_at[:10] if start_at else "start").replace("-", "")
        b = (end_at[:10] if end_at else "now").replace("-", "")
        suffix = f"-{a}-{b}"
    return f"drama-wordbook-{stamp}{suffix}.{ext.lstrip('.')}"


def content_disposition(filename: str) -> str:
    return f"attachment; filename=\"{filename}\"; filename*=UTF-8''{quote(filename)}"


def _meaning_text(it: dict) -> str:
    meanings = it.get("meanings") or []
    if isinstance(meanings, list):
        return "；".join(str(x) for x in meanings if str(x).strip())
    return str(meanings)


def _tag_text(value: object) -> str:
    if isinstance(value, list):
        return "、".join(str(x) for x in value if str(x).strip())
    return str(value or "")


def _ptext(value: object) -> str:
    return escape(str(value or ""), {"'": "&#39;", '"': "&quot;"})


def _item_pos(it: dict) -> str:
    text = str(it.get("dictionary_form") or it.get("surface") or "").strip()
    if not text:
        return ""
    try:
        toks = tokenize_ja(text, include_stop=True)
    except Exception:
        return ""
    for tok in toks:
        if str(tok.get("dictionary_form") or tok.get("surface") or "").strip() == text:
            return str(tok.get("pos") or "")
    return str((toks[0] if toks else {}).get("pos") or "")


def build_wordbook_xlsx_bytes(start: str | None = None, end: str | None = None) -> bytes:
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Font, PatternFill
    from openpyxl.utils import get_column_letter

    start_at, end_at = normalize_export_range(start, end)
    items = get_all_vocab_items(start_at=start_at, end_at=end_at)
    sentences = get_all_sentences_flat(start_at=start_at, end_at=end_at)

    wb = Workbook()
    ws_vocab = wb.active
    assert ws_vocab is not None
    ws_vocab.title = "生词"
    ws_sent = wb.create_sheet("句子")

    vocab_headers = [
        "id", "词条UUID", "词头ID", "词头（辞书形）", "表记", "读音", "音调", "词性", "JLPT", "释义（中文）",
        "例句（日）", "例句（中）", "标签", "来源", "句子ID", "句子UUID", "平台", "链接", "播放页标题", "剧集名",
        "分集名", "播放时间(s)", "截图路径", "创建时间", "更新时间",
    ]
    ws_vocab.append(vocab_headers)

    for it in items:
        plat, url, title, series, episode, tsec = _playback_columns(it.get("playback"))
        sid = it.get("sentence_id")
        ws_vocab.append([
            it.get("id"), it.get("uuid") or "", it.get("head_id"), it.get("dictionary_form") or "", it.get("surface") or "",
            it.get("reading") or "", "" if it.get("accent") is None else it.get("accent"), _item_pos(it), it.get("jlpt_level") or "",
            _meaning_text(it), it.get("example_ja") or "", it.get("example_zh") or "", _tag_text(it.get("tags")), it.get("source") or "",
            sid if sid is not None else "", it.get("sentence_uuid") or "", plat, url, title, series, episode, tsec,
            it.get("screenshot_path") or "", it.get("created_at") or "", it.get("updated_at") or "",
        ])

    sent_headers = [
        "id", "句子UUID", "日文", "中文", "标签", "关联词数", "来源", "平台", "链接", "播放页标题", "剧集名", "分集名",
        "播放时间(s)", "截图路径", "创建时间", "更新时间",
    ]
    ws_sent.append(sent_headers)

    for s in sentences:
        plat, url, title, series, episode, tsec = _playback_columns(s.get("playback"))
        ws_sent.append([
            s.get("id"), s.get("uuid") or "", s.get("example_ja") or "", s.get("example_zh") or "", _tag_text(s.get("tags")),
            s.get("word_count"), s.get("source") or "", plat, url, title, series, episode, tsec, s.get("screenshot_path") or "",
            s.get("created_at") or "", s.get("updated_at") or "",
        ])

    header_fill = PatternFill("solid", fgColor="2E8F76")
    header_font = Font(bold=True, color="FFFFFF")
    for ws in (ws_vocab, ws_sent):
        ws.freeze_panes = "A2"
        for cell in ws[1]:
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal="center", vertical="center")
        widths = [10, 20, 10, 18, 18, 18, 8, 12, 10, 28, 38, 32, 18, 10, 10, 20, 12, 28, 24, 20, 18, 12, 34, 24, 24]
        for idx, width in enumerate(widths[: ws.max_column], start=1):
            ws.column_dimensions[get_column_letter(idx)].width = width
        for row in ws.iter_rows(min_row=2):
            for cell in row:
                cell.alignment = Alignment(vertical="top", wrap_text=True)

    bio = BytesIO()
    wb.save(bio)
    return bio.getvalue()


def _episode_key(it: dict) -> tuple[str, str]:
    pb = it.get("playback") or {}
    series = str(pb.get("series_name") or pb.get("title") or "未命名剧集").strip() or "未命名剧集"
    episode = str(pb.get("episode_name") or "未命名分集").strip() or "未命名分集"
    return series, episode


def _accent_label(value: object) -> str:
    return "-" if value is None or value == "" else str(value)


def _shot_path(it: dict) -> str:
    path = str(it.get("screenshot_path") or "").strip()
    return path if path and Path(path).exists() else ""


def build_wordbook_pdf_bytes(start: str | None = None, end: str | None = None) -> bytes:
    from reportlab.lib import colors
    from reportlab.lib.enums import TA_CENTER
    from reportlab.lib.pagesizes import A4
    from reportlab.lib.styles import ParagraphStyle, getSampleStyleSheet
    from reportlab.lib.units import mm
    from reportlab.pdfbase import pdfmetrics
    from reportlab.pdfbase.cidfonts import UnicodeCIDFont
    from reportlab.pdfbase.ttfonts import TTFont
    from reportlab.platypus import Image, KeepTogether, PageBreak, Paragraph, SimpleDocTemplate, Spacer, Table, TableStyle

    start_at, end_at = normalize_export_range(start, end)
    items = get_all_vocab_items(start_at=start_at, end_at=end_at)

    for font_name in ("HeiseiMin-W3", "HeiseiKakuGo-W5", "STSong-Light"):
        try:
            pdfmetrics.registerFont(UnicodeCIDFont(font_name))
        except Exception:
            pass

    def register_ttf(name: str, candidates: list[str]) -> str | None:
        for candidate in candidates:
            path = Path(candidate)
            if not path.exists():
                continue
            try:
                pdfmetrics.registerFont(TTFont(name, str(path), subfontIndex=0))
                return name
            except Exception:
                continue
        return None

    cjk_regular = register_ttf(
        "DramaCJKRegular",
        [
            "C:/Windows/Fonts/msyh.ttc",
            "C:/Windows/Fonts/msyh.ttf",
            "C:/Windows/Fonts/simhei.ttf",
            "/System/Library/Fonts/PingFang.ttc",
            "/System/Library/Fonts/Hiragino Sans GB.ttc",
            "/Library/Fonts/Arial Unicode.ttf",
            "/System/Library/Fonts/Supplemental/Arial Unicode.ttf",
            "/usr/share/fonts/opentype/noto/NotoSansCJK-Regular.ttc",
            "/usr/share/fonts/opentype/noto/NotoSansCJKsc-Regular.otf",
        ],
    )
    cjk_bold = register_ttf(
        "DramaCJKBold",
        [
            "C:/Windows/Fonts/msyhbd.ttc",
            "C:/Windows/Fonts/msyhbd.ttf",
            "C:/Windows/Fonts/simhei.ttf",
            "/System/Library/Fonts/PingFang.ttc",
            "/System/Library/Fonts/Hiragino Sans GB.ttc",
            "/usr/share/fonts/opentype/noto/NotoSansCJK-Bold.ttc",
            "/usr/share/fonts/opentype/noto/NotoSansCJKsc-Bold.otf",
        ],
    )
    regular_font = cjk_regular or "HeiseiMin-W3"
    bold_font = cjk_bold or cjk_regular or "HeiseiKakuGo-W5"
    cn_font = cjk_regular or "STSong-Light"

    styles = getSampleStyleSheet()
    base = ParagraphStyle("BaseCjk", parent=styles["Normal"], fontName=regular_font, fontSize=9.5, leading=13, textColor=colors.HexColor("#243142"))
    cn = ParagraphStyle("Cn", parent=base, fontName=cn_font, fontSize=9.2, leading=12.5, textColor=colors.HexColor("#526173"))
    title = ParagraphStyle("TitleCjk", parent=base, fontName=bold_font, fontSize=22, leading=28, alignment=TA_CENTER, textColor=colors.HexColor("#1f2a37"))
    subtitle = ParagraphStyle("SubtitleCjk", parent=base, alignment=TA_CENTER, fontSize=10, textColor=colors.HexColor("#667085"))
    section = ParagraphStyle("SectionCjk", parent=base, fontName=bold_font, fontSize=14, leading=18, textColor=colors.white)
    word = ParagraphStyle("WordCjk", parent=base, fontName=bold_font, fontSize=13, leading=17, textColor=colors.HexColor("#1f2a37"))
    meta = ParagraphStyle("MetaCjk", parent=base, fontSize=8.5, leading=11, textColor=colors.HexColor("#667085"))
    example = ParagraphStyle("ExampleCjk", parent=base, fontSize=9.4, leading=13.5, textColor=colors.HexColor("#344054"))

    bio = BytesIO()
    doc = SimpleDocTemplate(
        bio,
        pagesize=A4,
        rightMargin=15 * mm,
        leftMargin=15 * mm,
        topMargin=16 * mm,
        bottomMargin=15 * mm,
        title="Drama Wordbook",
    )
    story = [Paragraph("Drama Wordbook", title)]
    range_text = "全部收藏"
    if start_at or end_at:
        range_text = f"{start_at[:10] if start_at else '开始'} - {end_at[:10] if end_at else '现在'}"
    story += [Paragraph(f"按剧集整理 · {range_text} · {len(items)} 个生词", subtitle), Spacer(1, 8 * mm)]

    if not items:
        story.append(Paragraph("这个时间段还没有可导出的生词。", cn))
        doc.build(story)
        return bio.getvalue()

    grouped: dict[tuple[str, str], list[dict]] = {}
    for it in items:
        grouped.setdefault(_episode_key(it), []).append(it)

    first_section = True
    for (series, episode), episode_items in grouped.items():
        if not first_section:
            story.append(PageBreak())
        first_section = False
        header = Table([[Paragraph(f"{_ptext(series)} / {_ptext(episode)}", section)]], colWidths=[180 * mm])
        header.setStyle(TableStyle([
            ("BACKGROUND", (0, 0), (-1, -1), colors.HexColor("#2E8F76")),
            ("BOX", (0, 0), (-1, -1), 0, colors.HexColor("#2E8F76")),
            ("LEFTPADDING", (0, 0), (-1, -1), 8),
            ("RIGHTPADDING", (0, 0), (-1, -1), 8),
            ("TOPPADDING", (0, 0), (-1, -1), 7),
            ("BOTTOMPADDING", (0, 0), (-1, -1), 7),
        ]))
        story.extend([header, Spacer(1, 4 * mm)])

        for it in episode_items:
            surface = str(it.get("surface") or it.get("dictionary_form") or "")
            dictionary = str(it.get("dictionary_form") or surface)
            reading = str(it.get("reading") or "-")
            pos = _item_pos(it) or "-"
            accent = _accent_label(it.get("accent"))
            meanings = _meaning_text(it) or "-"
            ja = str(it.get("example_ja") or "")
            zh = str(it.get("example_zh") or "")
            pb = it.get("playback") or {}
            time_text = "" if pb.get("current_time") is None else f" · {int(float(pb.get('current_time') or 0))}s"

            left = [
                Paragraph(f"{_ptext(surface)} <font color='#667085'>/ {_ptext(dictionary)}</font>", word),
                Paragraph(f"读音 {_ptext(reading)} · 音调 {_ptext(accent)} · 词性 {_ptext(pos)}{_ptext(time_text)}", meta),
                Paragraph(_ptext(meanings), cn),
                Spacer(1, 1.6 * mm),
                Paragraph(_ptext(ja or "-"), example),
                Paragraph(_ptext(zh or ""), cn),
            ]
            shot = _shot_path(it)
            if shot:
                try:
                    right = Image(shot, width=42 * mm, height=26 * mm, kind="proportional")
                except Exception:
                    right = Paragraph("暂无剧照", meta)
            else:
                right = Paragraph("暂无剧照", meta)

            card = Table([[left, right]], colWidths=[128 * mm, 46 * mm])
            card.setStyle(TableStyle([
                ("BOX", (0, 0), (-1, -1), 0.7, colors.HexColor("#D9E2EC")),
                ("BACKGROUND", (0, 0), (-1, -1), colors.HexColor("#FFFFFF")),
                ("VALIGN", (0, 0), (-1, -1), "TOP"),
                ("LEFTPADDING", (0, 0), (-1, -1), 8),
                ("RIGHTPADDING", (0, 0), (-1, -1), 8),
                ("TOPPADDING", (0, 0), (-1, -1), 7),
                ("BOTTOMPADDING", (0, 0), (-1, -1), 7),
            ]))
            story.extend([KeepTogether([card]), Spacer(1, 3 * mm)])

    doc.build(story)
    return bio.getvalue()
